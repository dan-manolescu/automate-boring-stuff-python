#! python3
# spreadsheetToText.py - Writes the cells of each column in the spreadsheet to a different text file.
# Usage: py spreadsheetToText.py <spreadsheet> <outputfilename>
# Output files are going to be outputfilename + counter + '.txt'
# Current limitations: only the active worksheet is being processed!

import openpyxl, sys, os

def spreadsheetToText(spreadsheet: str, outputFileName: str) -> None:
    '''
    Reads the contents of spreadsheet. For each column write the lines of text
    to a filename named outputfilename + column number + txt extension.
    '''
    if not os.path.exists(spreadsheet):
        print('No spreadsheet file found!')
        return

    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.active
    for col in range(1, sheet.max_column + 1):
        fileName = outputFileName + str(col) + '.txt'
        print(f'Writing column {col} to file {fileName}')

        file = open(fileName, 'w')

        for row in range(1, sheet.max_row + 1):
            cellVal = sheet.cell(row=row, column=col).value
            file.write(cellVal if cellVal != None else '')

        file.close()

    print('Done!')


if len(sys.argv) < 3:
    print(f'Usage: py {sys.argv[0]} spreadsheet outputfilename')
    sys.exit()

spreadsheetToText(sys.argv[1], sys.argv[2])
