#! python3
# spreadsheetInverter.py - Inverts the row and column of the cells in the spreadsheet.
# Usage: py spreadsheetInverter.py spreadsheet.

import openpyxl, sys, os

def invertSpreadsheet(spreadsheet: str) -> None:
    '''
    Inverts the row and column of the spreadsheet.
    '''
    wb = openpyxl.load_workbook(spreadsheet)

    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        # Make a new sheet where to copy the inverted cells
        newSheet = wb.create_sheet()

        # Copy the inverted cells into the new sheet
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                newSheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

        # Delete the original sheet and rename the new one
        del wb[sheetName]
        newSheet.title = sheetName

    wb.save(spreadsheet)

if len(sys.argv) < 2:
    print(f'Usage: py {sys.argv[0]} spreadsheet')
    sys.exit()

if not os.path.exists(sys.argv[1]):
    print('Invalid spreadsheet file!')
    sys.exit()

invertSpreadsheet(sys.argv[1])
