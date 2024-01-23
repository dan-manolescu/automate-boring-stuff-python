#! python3
# sendDuesReminders.py - Sends emails based on payment status it spreadsheet.

import openpyxl, ezgmail, sys

# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account.
ezgmail.init()
# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = '''Dear %s,
    Records show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!''' % (name, latestMonth)
    print('Sending email to %s...' % email)
    ezgmail.send(email, '%s dues unpaid' % latestMonth, body)
