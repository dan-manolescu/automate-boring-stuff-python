#! python3
# excelToCsv.py - Reads all Excel files in the current working directory and output them as CSV files.

import openpyxl, csv, os

for excelFile in os.listdir():
    if not excelFile.endswith('.xlsx'):
        continue  # skip non-xlsx files

    # Load the workbook object.
    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        print('Processing excel file', excelFile, '; sheet', sheetName)
        sheet = wb[sheetName]
        # Create the CSV filename from the Excel filename and sheet title.
        csvFilename = excelFile[:-5] + '_' + sheetName + '.csv'
        print('Writing to CSV file', csvFilename)
        csvFileObj = open(csvFilename, 'w', newline='')
        csvWriter = csv.writer(csvFileObj)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []  # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)

            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)

        csvFileObj.close()

